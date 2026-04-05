from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\jichu\Downloads\JOPETWIL71\JPT71-20260309T141921Z-1-001")
SOURCE_WORKBOOKS = [
    ROOT / "JPT.xlsx",
    ROOT / "JPT71" / "JPT.xlsx",
]
RECON_WORKBOOKS = [
    ROOT / "JPT-reconciled_v6.0.xlsx",
    ROOT / "dashboard_final" / "JPT-reconciled_v6.0.xlsx",
    ROOT / "JPT71_AI_Team_Share_Easy" / "dashboard" / "JPT-reconciled_v6.0.xlsx",
    ROOT / "JPT71_Dashboard_Final_Email_Pack_2026-03-10" / "JPT-reconciled_v6.0.xlsx",
    ROOT / "output" / "pack-staging" / "JPT71_AI_Team_Share_Easy" / "dashboard" / "JPT-reconciled_v6.0.xlsx",
    ROOT / "output" / "share-pack-staging-20260310" / "easy" / "dashboard" / "JPT-reconciled_v6.0.xlsx",
]
SYNTHETIC_MARKER = "MISSING_VDR_22-FEB-2026_INSERTED_FROM_98256402"
TARGET_INVOICE = "98256402"


@dataclass
class CheckResult:
    path: Path
    ok: bool
    detail: str


def count_decklog_rows(ws, invoice_col: int, source_col: int) -> tuple[int, int]:
    matched = 0
    synthetic = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        dt = row[1]
        vessel = str(row[2] or "").strip().upper()
        invoice = str(row[invoice_col - 1] or "").strip()
        source = str(row[source_col - 1] or "").strip()
        if "JOPETWIL" in vessel and getattr(dt, "year", None) == 2026 and getattr(dt, "month", None) == 2:
            if invoice == TARGET_INVOICE:
                matched += 1
        if source == SYNTHETIC_MARKER:
            synthetic += 1
    return matched, synthetic


def check_source_workbook(path: Path) -> CheckResult:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["DECKLOG"]
    matched, synthetic = count_decklog_rows(ws, 33, 32)
    ok = matched == 27 and synthetic == 0
    detail = f"decklog_rows={matched}, synthetic_rows={synthetic}"
    return CheckResult(path, ok, detail)


def check_recon_workbook(path: Path) -> CheckResult:
    wb = load_workbook(path, read_only=True, data_only=True)

    decklog_ws = wb["1_Decklog"]
    matched, synthetic = count_decklog_rows(decklog_ws, 33, 32)

    ctx_row = None
    for row in wb["8_Decklog_Context"].iter_rows(min_row=2, values_only=True):
        if str(row[1] or "").strip() == "2026-02" and str(row[2] or "").strip().upper() == "JOPETWIL71":
            ctx_row = row
            break

    progress_found = False
    for row in wb["9_ProgressPay_Inv"].iter_rows(min_row=2, values_only=True):
        if str(row[2] or "").strip() == TARGET_INVOICE:
            progress_found = True
            break

    als_rows = sum(
        1
        for row in wb["11_ALS_Allocation"].iter_rows(min_row=2, values_only=True)
        if str(row[0] or "").strip() == TARGET_INVOICE
    )
    cost_rows = sum(
        1
        for row in wb["12_Cost_Summary"].iter_rows(min_row=2, values_only=True)
        if str(row[0] or "").strip()
        in {"HVDC-AGI-GRM-J71-094", "HVDC-AGI-BIN-J71-095", "HVDC-AGI-BIN-J71-096", "HVDC-AGI-BIN-J71-097"}
    )

    ctx_ok = (
        ctx_row is not None
        and str(ctx_row[0]).strip() == TARGET_INVOICE
        and int(ctx_row[4]) == 27
    )

    ok = matched == 27 and synthetic == 0 and ctx_ok and progress_found and als_rows == 0 and cost_rows == 0
    detail = (
        f"decklog_rows={matched}, synthetic_rows={synthetic}, "
        f"context_ok={ctx_ok}, progress_found={progress_found}, als_rows={als_rows}, cost_rows={cost_rows}"
    )
    return CheckResult(path, ok, detail)


def main() -> None:
    results: list[CheckResult] = []
    for path in SOURCE_WORKBOOKS:
        results.append(check_source_workbook(path))
    for path in RECON_WORKBOOKS:
        results.append(check_recon_workbook(path))

    failed = [r for r in results if not r.ok]
    for result in results:
        status = "OK" if result.ok else "FAIL"
        print(f"{status} {result.path}: {result.detail}")

    if failed:
        raise SystemExit(1)


if __name__ == "__main__":
    main()

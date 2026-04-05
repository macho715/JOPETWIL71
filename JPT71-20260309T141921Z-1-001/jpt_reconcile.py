from __future__ import annotations

import math
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(r"C:\Users\minky\Downloads")
SOURCE_PATH = BASE_DIR / "JPT.xlsx"
OUTPUT_DIR = BASE_DIR / "artifacts"
OUTPUT_PATH = OUTPUT_DIR / "JPT-reconciled.xlsx"

VOYAGE_SHEET = "VOYAGE"
OFCO_SHEET = "OFCO INV"
DECKLOG_SHEET = "DECKLOG"


def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ").strip()
    if not text:
        return ""
    return re.sub(r"\s+", " ", text)


def normalize_text(value) -> str:
    return clean_text(value).upper()


def normalize_voyage_id(value) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if "-J71-" in text:
        return text
    return text.replace("-71-", "-J71-")


def normalize_invoice(value) -> str:
    return normalize_text(value)


def safe_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)

    text = clean_text(value)
    if not text:
        return 0.0
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def normalize_date(value):
    if isinstance(value, datetime):
        return value
    return value


def format_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = clean_text(value)
    return text


def unique_join(values) -> str:
    ordered = []
    seen = set()
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        key = text.upper()
        if key in seen:
            continue
        seen.add(key)
        ordered.append(text)
    return " | ".join(ordered)


def header_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        header = clean_text(cell.value)
        if header and header not in mapping:
            mapping[header] = idx
    return mapping


def cell_value(ws, row_idx: int, headers: dict[str, int], header: str):
    col = headers[clean_text(header)]
    return ws.cell(row_idx, col).value


def append_rows(ws, rows: list[list]):
    for row in rows:
        ws.append(row)


def style_table(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_len = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 40)


def voyage_rollup(wb):
    ws = wb[VOYAGE_SHEET]
    headers = header_map(ws)
    groups = {}

    for row_idx in range(2, ws.max_row + 1):
        original_voyage = cell_value(ws, row_idx, headers, "Voyage No")
        normalized_voyage = normalize_voyage_id(original_voyage)
        if not normalized_voyage:
            continue

        group = groups.setdefault(
            normalized_voyage,
            {
                "original_voyage_ids": [],
                "batches": [],
                "loading_dates": [],
                "sub_cons": [],
                "descriptions": [],
                "sizes": [],
                "delivery_qty": 0.0,
                "invoice_numbers": [],
                "row_count": 0,
            },
        )
        group["original_voyage_ids"].append(original_voyage)
        group["batches"].append(cell_value(ws, row_idx, headers, "Batch"))
        loading_date = normalize_date(cell_value(ws, row_idx, headers, "Loading Date"))
        if loading_date:
            group["loading_dates"].append(loading_date)
        group["sub_cons"].append(cell_value(ws, row_idx, headers, "Sub-Con"))
        group["descriptions"].append(cell_value(ws, row_idx, headers, "Description\nof Material"))
        group["sizes"].append(cell_value(ws, row_idx, headers, "Size(mm or bag)"))
        group["delivery_qty"] += safe_float(cell_value(ws, row_idx, headers, "Delivery Qty.\nin Ton"))
        group["invoice_numbers"].append(cell_value(ws, row_idx, headers, "OFCO INVOICE NO"))
        group["row_count"] += 1

    output_rows = [
        [
            "Normalized Voyage ID",
            "Original Voyage IDs",
            "Batch Values",
            "Sub-Con Values",
            "Loading Date Start",
            "Loading Date End",
            "Descriptions",
            "Sizes",
            "Delivery Qty (Ton)",
            "VOYAGE Invoice Numbers",
            "Source Row Count",
        ]
    ]
    for key in sorted(groups):
        group = groups[key]
        dates = sorted(d for d in group["loading_dates"] if isinstance(d, datetime))
        output_rows.append(
            [
                key,
                unique_join(group["original_voyage_ids"]),
                unique_join(group["batches"]),
                unique_join(group["sub_cons"]),
                format_date(dates[0]) if dates else "",
                format_date(dates[-1]) if dates else "",
                unique_join(group["descriptions"]),
                unique_join(group["sizes"]),
                round(group["delivery_qty"], 2),
                unique_join(group["invoice_numbers"]),
                group["row_count"],
            ]
        )

    return groups, output_rows


def ofco_rollup(wb):
    ws = wb[OFCO_SHEET]
    headers = header_map(ws)
    groups = {}

    for row_idx in range(2, ws.max_row + 1):
        original_voyage = cell_value(ws, row_idx, headers, "Voyage No")
        invoice_number = cell_value(ws, row_idx, headers, "INVOICE NUMBER")
        normalized_voyage = normalize_voyage_id(original_voyage)
        normalized_invoice = normalize_invoice(invoice_number)
        if not normalized_voyage:
            continue

        key = (normalized_voyage, normalized_invoice)
        group = groups.setdefault(
            key,
            {
                "original_voyage_ids": [],
                "original_invoice_numbers": [],
                "subjects": [],
                "year_months": [],
                "cost_mains": [],
                "cost_center_as": [],
                "cost_center_bs": [],
                "price_centers": [],
                "total_aed": 0.0,
                "amount_usd": 0.0,
                "vat_usd": 0.0,
                "total_usd": 0.0,
                "line_count": 0,
            },
        )
        group["original_voyage_ids"].append(original_voyage)
        group["original_invoice_numbers"].append(invoice_number)
        group["subjects"].append(cell_value(ws, row_idx, headers, "SUBJECT"))
        group["year_months"].append(cell_value(ws, row_idx, headers, "YEAR_MONTH"))
        group["cost_mains"].append(cell_value(ws, row_idx, headers, "COST MAIN"))
        group["cost_center_as"].append(cell_value(ws, row_idx, headers, "COST CENTER A"))
        group["cost_center_bs"].append(cell_value(ws, row_idx, headers, "COST CENTER B"))
        group["price_centers"].append(cell_value(ws, row_idx, headers, "PRICE CENTER"))
        group["total_aed"] += safe_float(cell_value(ws, row_idx, headers, "Total AED"))
        group["amount_usd"] += safe_float(cell_value(ws, row_idx, headers, "Amount USD"))
        group["vat_usd"] += safe_float(cell_value(ws, row_idx, headers, "VAT USD"))
        group["total_usd"] += safe_float(cell_value(ws, row_idx, headers, "Total USD"))
        group["line_count"] += 1

    output_rows = [
        [
            "Normalized Voyage ID",
            "Normalized Invoice Number",
            "Original Voyage IDs",
            "Original Invoice Numbers",
            "Year-Month Values",
            "Subjects",
            "Cost Main Values",
            "Cost Center A Values",
            "Cost Center B Values",
            "Price Center Values",
            "Line Count",
            "Total AED",
            "Amount USD",
            "VAT USD",
            "Total USD",
        ]
    ]

    by_voyage = defaultdict(
        lambda: {
            "invoice_numbers": [],
            "total_aed": 0.0,
            "amount_usd": 0.0,
            "vat_usd": 0.0,
            "total_usd": 0.0,
            "line_count": 0,
            "group_count": 0,
            "subjects": [],
        }
    )

    for (voyage_key, invoice_key) in sorted(groups):
        group = groups[(voyage_key, invoice_key)]
        output_rows.append(
            [
                voyage_key,
                invoice_key,
                unique_join(group["original_voyage_ids"]),
                unique_join(group["original_invoice_numbers"]),
                unique_join(group["year_months"]),
                unique_join(group["subjects"]),
                unique_join(group["cost_mains"]),
                unique_join(group["cost_center_as"]),
                unique_join(group["cost_center_bs"]),
                unique_join(group["price_centers"]),
                group["line_count"],
                round(group["total_aed"], 2),
                round(group["amount_usd"], 2),
                round(group["vat_usd"], 2),
                round(group["total_usd"], 2),
            ]
        )

        voyage_group = by_voyage[voyage_key]
        voyage_group["invoice_numbers"].append(invoice_key)
        voyage_group["total_aed"] += group["total_aed"]
        voyage_group["amount_usd"] += group["amount_usd"]
        voyage_group["vat_usd"] += group["vat_usd"]
        voyage_group["total_usd"] += group["total_usd"]
        voyage_group["line_count"] += group["line_count"]
        voyage_group["group_count"] += 1
        voyage_group["subjects"].append(unique_join(group["subjects"]))

    return groups, by_voyage, output_rows


def build_summary(voyage_groups, ofco_by_voyage):
    all_keys = sorted(set(voyage_groups) | set(ofco_by_voyage))
    output_rows = [
        [
            "Normalized Voyage ID",
            "VOYAGE Present",
            "OFCO Present",
            "VOYAGE Original IDs",
            "Loading Date Start",
            "Loading Date End",
            "VOYAGE Descriptions",
            "VOYAGE Delivery Qty (Ton)",
            "VOYAGE Invoice Numbers",
            "OFCO Invoice Numbers",
            "Invoice Mismatch",
            "OFCO Invoice Group Count",
            "OFCO Line Count",
            "OFCO Total AED",
            "OFCO Amount USD",
            "OFCO VAT USD",
            "OFCO Total USD",
            "Match Status",
            "Notes",
        ]
    ]

    for key in all_keys:
        voyage_group = voyage_groups.get(key)
        ofco_group = ofco_by_voyage.get(key)
        voyage_present = voyage_group is not None
        ofco_present = ofco_group is not None

        voyage_invoices = set()
        if voyage_group:
            voyage_invoices = {
                normalize_invoice(value)
                for value in voyage_group["invoice_numbers"]
                if normalize_invoice(value)
            }

        ofco_invoices = set()
        if ofco_group:
            ofco_invoices = {
                normalize_invoice(value)
                for value in ofco_group["invoice_numbers"]
                if normalize_invoice(value)
            }

        invoice_mismatch = voyage_present and ofco_present and voyage_invoices != ofco_invoices
        if not voyage_present:
            status = "Missing in VOYAGE"
            notes = "Present only in OFCO invoice rollup."
        elif not ofco_present:
            status = "Missing in OFCO"
            notes = "Present only in VOYAGE rollup."
        elif invoice_mismatch:
            status = "Invoice Mismatch"
            notes = "VOYAGE and OFCO invoice lists differ after normalization."
        else:
            status = "Matched"
            notes = ""

        dates = []
        if voyage_group:
            dates = sorted(d for d in voyage_group["loading_dates"] if isinstance(d, datetime))

        output_rows.append(
            [
                key,
                "Yes" if voyage_present else "No",
                "Yes" if ofco_present else "No",
                unique_join(voyage_group["original_voyage_ids"]) if voyage_group else "",
                format_date(dates[0]) if dates else "",
                format_date(dates[-1]) if dates else "",
                unique_join(voyage_group["descriptions"]) if voyage_group else "",
                round(voyage_group["delivery_qty"], 2) if voyage_group else 0,
                unique_join(sorted(voyage_invoices)) if voyage_invoices else "",
                unique_join(sorted(ofco_invoices)) if ofco_invoices else "",
                "Yes" if invoice_mismatch else "No",
                ofco_group["group_count"] if ofco_group else 0,
                ofco_group["line_count"] if ofco_group else 0,
                round(ofco_group["total_aed"], 2) if ofco_group else 0,
                round(ofco_group["amount_usd"], 2) if ofco_group else 0,
                round(ofco_group["vat_usd"], 2) if ofco_group else 0,
                round(ofco_group["total_usd"], 2) if ofco_group else 0,
                status,
                notes,
            ]
        )

    return output_rows


def build_exceptions(summary_rows):
    output_rows = [summary_rows[0]]
    for row in summary_rows[1:]:
        if row[17] != "Matched":
            output_rows.append(row)
    return output_rows


def decklog_context(wb):
    ws = wb[DECKLOG_SHEET]
    headers = header_map(ws)
    groups = {}

    for row_idx in range(2, ws.max_row + 1):
        invoice_number = cell_value(ws, row_idx, headers, "Invoice No")
        date_value = normalize_date(cell_value(ws, row_idx, headers, "Date"))
        vessel = cell_value(ws, row_idx, headers, "Vessel")
        source_page = cell_value(ws, row_idx, headers, "Source_Page")
        normalized_invoice = normalize_invoice(invoice_number) or "[BLANK]"
        year_month = date_value.strftime("%Y-%m") if isinstance(date_value, datetime) else ""
        key = (normalized_invoice, year_month, clean_text(vessel))

        group = groups.setdefault(
            key,
            {
                "original_invoice_numbers": [],
                "dates": [],
                "vessels": [],
                "charterers": [],
                "call_ports": [],
                "source_pages": [],
                "row_count": 0,
            },
        )
        group["original_invoice_numbers"].append(invoice_number)
        if date_value:
            group["dates"].append(date_value)
        group["vessels"].append(vessel)
        group["charterers"].append(cell_value(ws, row_idx, headers, "Charterer"))
        group["call_ports"].append(cell_value(ws, row_idx, headers, "Call_to_Port"))
        group["source_pages"].append(source_page)
        group["row_count"] += 1

    output_rows = [
        [
            "Normalized Invoice Number",
            "Year-Month",
            "Vessel",
            "Original Invoice Numbers",
            "Row Count",
            "First Date",
            "Last Date",
            "Source Page Min",
            "Source Page Max",
            "Charterer Values",
            "Call to Port Values",
        ]
    ]

    for key in sorted(groups):
        group = groups[key]
        dates = sorted(d for d in group["dates"] if isinstance(d, datetime))
        pages = [safe_float(page) for page in group["source_pages"] if clean_text(page)]
        output_rows.append(
            [
                key[0],
                key[1],
                key[2],
                unique_join(group["original_invoice_numbers"]),
                group["row_count"],
                format_date(dates[0]) if dates else "",
                format_date(dates[-1]) if dates else "",
                int(min(pages)) if pages else "",
                int(max(pages)) if pages else "",
                unique_join(group["charterers"]),
                unique_join(group["call_ports"]),
            ]
        )

    return output_rows


def replace_or_create_sheet(wb, name: str):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(title=name)


def main():
    if not SOURCE_PATH.exists():
        raise FileNotFoundError(f"Source workbook not found: {SOURCE_PATH}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(SOURCE_PATH)

    voyage_groups, voyage_rows = voyage_rollup(wb)
    _, ofco_by_voyage, ofco_rows = ofco_rollup(wb)
    summary_rows = build_summary(voyage_groups, ofco_by_voyage)
    exception_rows = build_exceptions(summary_rows)
    decklog_rows = decklog_context(wb)

    outputs = {
        "VOYAGE Rollup": voyage_rows,
        "OFCO Rollup": ofco_rows,
        "Reconciliation Summary": summary_rows,
        "Exceptions": exception_rows,
        "DECKLOG Context": decklog_rows,
    }

    for sheet_name, rows in outputs.items():
        sheet = replace_or_create_sheet(wb, sheet_name)
        append_rows(sheet, rows)
        style_table(sheet)

    wb.save(OUTPUT_PATH)

    summary_status_counts = defaultdict(int)
    for row in summary_rows[1:]:
        summary_status_counts[row[17]] += 1

    print(f"Saved: {OUTPUT_PATH}")
    print(f"VOYAGE rollup rows: {len(voyage_rows) - 1}")
    print(f"OFCO rollup rows: {len(ofco_rows) - 1}")
    print(f"Summary rows: {len(summary_rows) - 1}")
    print(f"Exception rows: {len(exception_rows) - 1}")
    print("Statuses:")
    for status in sorted(summary_status_counts):
        print(f"  {status}: {summary_status_counts[status]}")


if __name__ == "__main__":
    main()
